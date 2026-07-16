from pathlib import Path

from langchain_community.document_loaders import (
    PyPDFLoader,
    TextLoader,
)
from langchain_core.documents import Document
from openpyxl import load_workbook


SUPPORTED_EXTENSIONS = {".pdf", ".txt", ".md", ".xlsx"}
VALID_CATEGORIES = {
    "general",
    "admission",
    "program",
    "campus",
    "pedagogy",
    "career",
    "event",
    "practical",
}


def detect_category(
    file_path: Path,
    documents_root: Path,
) -> str:
    try:
        relative_path = file_path.relative_to(
            documents_root
        )

    except ValueError:
        return "general"

    if len(relative_path.parts) < 2:
        return "general"

    category = (
        relative_path.parts[0]
        .strip()
        .lower()
    )

    if category not in VALID_CATEGORIES:
        return "general"

    return category

def load_pdf(file_path: Path) -> list[Document]:
    loader = PyPDFLoader(str(file_path))
    documents = loader.load()

    for document in documents:
        document.metadata.update(
            {
                "source": str(file_path),
                "file_name": file_path.name,
                "file_type": "pdf",
            }
        )

    return documents


def load_text(file_path: Path) -> list[Document]:
    loader = TextLoader(
        str(file_path),
        encoding="utf-8",
        autodetect_encoding=True,
    )

    documents = loader.load()

    for document in documents:
        document.metadata.update(
            {
                "source": str(file_path),
                "file_name": file_path.name,
                "file_type": file_path.suffix.removeprefix("."),
            }
        )

    return documents


def load_excel(file_path: Path) -> list[Document]:
    workbook = load_workbook(
        filename=file_path,
        read_only=True,
        data_only=True,
    )

    documents: list[Document] = []

    for sheet in workbook.worksheets:
        rows: list[str] = []

        for row_number, row in enumerate(
            sheet.iter_rows(values_only=True),
            start=1,
        ):
            values = [
                str(value).strip()
                for value in row
                if value is not None and str(value).strip()
            ]

            if not values:
                continue

            row_content = " | ".join(values)
            rows.append(f"Ligne {row_number}: {row_content}")

        if not rows:
            continue

        content = "\n".join(rows)

        document = Document(
            page_content=content,
            metadata={
                "source": str(file_path),
                "file_name": file_path.name,
                "file_type": "xlsx",
                "sheet_name": sheet.title,
            },
        )

        documents.append(document)

    workbook.close()

    return documents


def load_file(file_path: Path) -> list[Document]:
    extension = file_path.suffix.lower()

    if extension == ".pdf":
        return load_pdf(file_path)

    if extension in {".txt", ".md"}:
        return load_text(file_path)

    if extension == ".xlsx":
        return load_excel(file_path)

    raise ValueError(
        f"Format non pris en charge : {file_path.suffix}"
    )


def load_documents(directory: Path) -> list[Document]:
    if not directory.exists():
        raise FileNotFoundError(
            f"Le dossier des documents n'existe pas : {directory}"
        )

    if not directory.is_dir():
        raise NotADirectoryError(
            f"Le chemin n'est pas un dossier : {directory}"
        )

    documents: list[Document] = []

    files = sorted(
        file_path
        for file_path in directory.rglob("*")
        if file_path.is_file()
        and file_path.suffix.lower() in SUPPORTED_EXTENSIONS
    )

    for file_path in files:
        try:
            loaded_documents = load_file(file_path)

            category = detect_category(
                file_path=file_path,
                documents_root=directory,
            )

            for document in loaded_documents:
                document.metadata.update(
                    {
                        "category": category,
                        "source_type": (
                            "official_document"
                        ),
                    }
                )

            documents.extend(loaded_documents)

            print(
                f"[OK] {file_path.name}: "
                f"{len(loaded_documents)} document(s)"
            )

        except Exception as error:
            print(
                f"[ERREUR] Impossible de charger "
                f"{file_path.name}: {error}"
            )

    return documents